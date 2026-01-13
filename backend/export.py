from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from datetime import date
import openpyxl
from openpyxl.styles import Font, Alignment
from fastapi.responses import StreamingResponse
from io import BytesIO


from database import get_db
from services.dashboard import get_kabag_dashboard_data


router = APIRouter(prefix="/export", tags=["Export"])


@router.get("/ekinerja")
def export_ekinerja(period: str = "monthly", db: Session = Depends(get_db)):
data = get_kabag_dashboard_data(period, db)


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "e-Kinerja IT"


ws.merge_cells('A1:D1')
ws['A1'] = "LAPORAN e-KINERJA IT RSUD"
ws['A1'].font = Font(bold=True)
ws['A1'].alignment = Alignment(horizontal="center")


ws.append([])
ws.append(["Periode", period.upper()])
ws.append(["Tanggal Cetak", date.today().isoformat()])
ws.append([])


ws.append(["Total Tiket", data['total_tickets']])
ws.append(["Selesai Tepat Waktu", data['on_time']])
ws.append(["Terlambat", data['late']])
ws.append(["SLA Compliance (%)", data['sla_compliance']])


ws.append([])
ws.append(["Kategori", "Total", "On Time", "Late"])
for c in data['by_category']:
ws.append([c['category'], c['total'], c['on_time'], c['late']])


for col in ws.columns:
ws.column_dimensions[col[0].column_letter].width = 20


stream = BytesIO()
wb.save(stream)
stream.seek(0)


filename = f"e-kinerja-{period}.xlsx"
return StreamingResponse(stream,
media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
headers={"Content-Disposition": f"attachment; filename={filename}"}
)
